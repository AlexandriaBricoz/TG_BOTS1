import os

from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter


from io import BytesIO

def create_excel_table_in_memory(data):
    wb = Workbook()
    ws = wb.active

    headers = ["Имя", "Количество", "День", "Ночь", "Количество замен", "Остаток баланса"]

    ws.append(headers)

    for row in data:
        ws.append(row)

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20

    for row in ws.iter_rows(min_row=1, max_row=len(data)+1, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.alignment = Alignment(horizontal='center')

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer